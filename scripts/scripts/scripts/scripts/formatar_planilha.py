from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def formatar_planilha():
    print("Formatando planilha...")
    caminho = "../output/vendas_processado.xlsx"
    wb = load_workbook(caminho)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    wb.save(caminho)
    print("Planilha formatada.")
