from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def gerar_grafico():
    print("Gerando gráfico no Excel...")
    caminho = "../output/vendas_processado.xlsx"

    wb = load_workbook(caminho)
    ws = wb["Resumo"]

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Vendas por Região"
    chart.y_axis.title = 'Total (R$)'
    chart.x_axis.title = 'Região'

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "E2")
    wb.save(caminho)
    print("Gráfico gerado na aba Resumo.")
